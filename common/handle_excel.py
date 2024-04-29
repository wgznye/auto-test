
import openpyxl
from openpyxl.styles import Font
import time


timestamp = int(time.time())
casesId = "test:cases:" + str(timestamp)
class HandleExcel:
    def __init__(self, filename):
        self.filename = filename

    def read_file(self):
        """
        按行读取excel用例
        :return:
        """
        workbook = openpyxl.load_workbook(self.filename)
        cases = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            res = list(sheet.rows)
            # 取第一行表头
            title = [i.value for i in res[0]]
            # 遍历后续内容，与title打包传入列表
            for item in res[1:]:
                data = [i.value for i in item]
                data.append(sheet_name)
                print("data>>>>" ,data)
                cases.append(data)
        return cases

    def write_file(self, row, column, value, color=None):
        """
        回写excel
        :param row: 行
        :param column: 列
        :param value: 值
        :return:
        """
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheet]
        sh.cell(row=row, column=column, value=value).font = Font(color=color)
        workbook.save(self.filename)
